#合計が100点超えたら〇、そうでないなら×を転記
#100時間超えたフラグ判定の関数　returneで戻り値を返す

import openpyxl
wbPath = "../learning/kiso001_kaito.xlsx"
svPath = "../learning/kiso001_kaito_mod3.xlsx"
wb = openpyxl.load_workbook(wbPath)
ws = wb["Mondai1"]

#100時間超えたフラグ判定の関数　returneで戻り値を返す
def Hantei(overtime):
    if overtime >= 100:
        over = "〇"
    else:
        over = "×"
    return over

#リストに保存する　rowの実行結果　(<Cell 'Mondai1'.B2>, <Cell 'Mondai1'.C2>)(<Cell 'Mondai1'.B3>, <Cell 'Mondai1'.C3>)
ls = []
for row in ws["B2:C11"]:
    ls.append(row)
    #print(row)

#リストに保存する　row(1行分）をさらにfor文でまわして、ループをぬけたとき実行結果
#['米子 美和', 227] ['呉 早希', 65] ['五日市 加奈子', 139]
lst_total = []
for row in ws["B2:C11"]:
    lst = []#空リスト
    for r in row:
        lst.append(r.value)
    lst_total.append(lst)# インデントを一段あげる　ループをぬけたら、リストの要素0 ['米子 美和', 227] リストの要素１ ['呉 早希', 65]と保存
    #print(lst)

#名前、合計をセットにしたリストをfor文で調査、関数 def Hantei（引数：合計時間）を渡し、判定結果の戻り値をもらう
#書き出す行をcntのカウンターでカウントアップ　
#インデックス番号と値を取り出すenumerate関数もある。 for c,v in enumerate(lst_total):
cnt = 2
for c in lst_total:
    ans = Hantei(c[1])#合計時間
    print(ans,c[0],c[1],sep=":")#引数　リスト[要素の番号]
    ws["A"+ str(cnt)].value = ans#セルの番地は文字列str関数で文字列に変換
    cnt = cnt + 1

wb.save(svPath)