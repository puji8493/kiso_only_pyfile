#[3]シート「Sheet2」で、1～6月の残業時間が一番多い人の氏名をセルK4に、月名をセルL4に、残業時間の値をセルM4に記入するマクロを作りなさい。（ただし、残業時間が一番多い人はひとりしかいないものとしてよい）
#一番得点の高い人はを重複させたデータ　40.5　二名　40.0　1名いるとする
#最大値が２つ以上あった場合を考慮し、最大値のインデックス番号をリストに格納する
#最大値が何回目のループ回数か調べたいので、enumerate関数を使ってインデックスを抽出　#抽出したインデックス番号をlist.appned()メソッドで格納。appendメソッドつけ忘れて取り扱えないオブジェクトエラー注意

#最大値をfloat型で返すには、numpyを利用する
#インストールはコマンドプロンプトでpip install numpy　ＤＬしてから、import numpy 分かりやすい参考リンクhttps://udemy.benesse.co.jp/data-science/ai/python-numpy.html
# arry の公式ドキュメント　https://docs.python.org/ja/3/library/array.html
#numpyの配列の取り扱い　分かりやすかったリンク　https://python.atelierkobato.com/max/
import openpyxl
import numpy as np
ph = "../5bai/ks012.xlsx"
mod_ph = "../5bai/ks012_mod3.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet2"]

month_list = ["1月","2月","3月","4月","5月","6月"]
ls = []#row1行の時間を格納するリスト
v_ls = [] #値だけのリスト　max関数を使う
cnt = 6
#数字の範囲だけ繰り返し処理　
for row in sh["C6:H33"]:
    #B列氏名の変数は、行をカンター変数を文字列に変換してセル番地で指定
    namae = sh["B"+str(cnt)].value
    for c,cl in enumerate(row):
        month = month_list[c]
        ls.append([namae,month,cl.value])
        v_ls.append(cl.value)
    cnt = cnt + 1

#最大値を変数に格納
max_value = max(v_ls)
print(max_value)    

#残業時間表の中でmax関数の戻り値と一致したら、セルＫ４から書き出す
#最大値が複数あったら全て書き出す
cnt = 4
for c,v in enumerate(ls):
    if max_value == v[2]:
        print(c,v,sep=":")
        print(v[0])
        sh.cell(row=cnt,column=11).value = v[0]
        sh.cell(row=cnt,column=12).value = v[1]
        sh.cell(row=cnt,column=13).value = v[2]
        cnt = cnt + 1

wb.save(mod_ph)