##100時間残業時間超過の人を要注意リストに格納して、書き出す
import openpyxl
ph = "../learning/ks004.xlsx"
mod_ph = "../learning/ks004_mod2.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet1"]

#データが入力されているセルの最大行は「max_row」最大列は「max_column」で取得
max_gyou = sh.max_row
print(max_gyou)

#全員の残業リスト
zangyo_list = []
for row in sh.iter_rows(min_row=8,max_row=max_gyou,max_col=3):
    #行の空リスト
    lst = []
    for r in row:
        lst.append(r.value)
    zangyo_list.append(lst)
    #print(lst)
    #実行結果
    #[1, '五日市 加奈子', 139]
    #[2, '呉 早希', 65]
    #[3, '高岡 範夫', 0]

#全員の残業リストから、100時間以上の人を要注意リストに格納する
#「リスト内包表記」で、ループ処理を１行に繋げて書くことができます。
#100時間以上の残業対象者をリスト内包表記でリストに格納する
# [仮変数を用いた式 for 仮変数 in イテラブルオブジェクト（反復可能なオブジェクト）if 条件式]
a_list = [t for t in zangyo_list if t[2]>=100]
print(a_list)

for c,v in enumerate(a_list):
    #要注意リストの要素を書き出す　v[0]=ID　v[1]=名前 v[2]=合計
    sh.cell(row=5+c,column=6).value = v[0]
    sh.cell(row=5+c,column=7).value = v[1]
    sh.cell(row=5+c,column=8).value = v[2]

wb.save(mod_ph)