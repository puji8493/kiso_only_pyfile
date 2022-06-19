#[1]当月支払金額のリストを辞書で格納
#[2]辞書　同一のkeyの値は累計額を加算する　結婚お祝いなら、10,000+5000=15,000
#   dicのキーが存在しない場合　if key not in dic.keys() 他にdic.get(key)
#   dicの値を取り出すのは　dic[key] dic["結婚お祝い"] + spend_list[c1]
#[3] 辞書のキーと値を取り出すのは、 for k,v in dic.items()  #実行結果　結婚祝い 15000　お見舞い袋 100　出産祝い 10000　御霊前 5000　お見舞い 3000
#[4] 辞書型（dictionary）の値をvaluesでリスト型にして、まとめて取り出す dic.values()　
# 実行結果dict_values([15000, 100, 10000, 5000, 3000])
#　辞書のアイテムをsum関数で足すには　sum(dic.value()のリスト)
#[5] 総覧表　D,E列に書き出し まずはリストに格納
#[6] 支払総計表への転記
# リストの値 = リスト型　class List
# リストの要素を取り出す　str型　リスト[0] = str
# 辞書のvalueを比較する場合、同じstr型出ないと一致しない　よって、リスト[0]（添え字）と辞書の値を比較する
# 支払累計額のリストを、ruikei_listで作成　最後に総類額をsum(ruikei_list.values())で取り出して転記
#リストの要素が、4月支払リストの辞書のキーにない場合は、先月までの支払総額を、支払累計額に転記する
#   continueでそれ以降を中断　ループを続行
#リストの要素が、4月支払リストの辞書のキーにある場合は、4月のKeyの値を、今月支払額に転記する
#   今月支払額は、4月支払リストの辞書の値、支払累計学は、sum([先月までの支払総額,4月支出の辞書の値])
#   breakでループを終了
#[7]今月支払額の累計　sum(4月支払リストの辞書.values())
#   辞書型（dictionary）の値をvaluesでリスト型にして、まとめて取り出す sum(list名)
#   支払累計額 リストの累計額は、sum([list])

import os
import openpyxl
print(os.getcwd())
ph = "../learning/ks011.xlsx"
mod_ph = "../learning/ks011_mod06.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet1"]

#[1]当月支払金額のリストを辞書で格納
dic_spend = {}
spend_list = []
for row in sh["I4:J10"]:
    row_list = []
    for r in row:
        row_list.append(r.value)
    spend_list.append(row_list)
#print(spend_list)
#実行結果　[['結婚祝い', 10000], ['お見舞い袋', 100], ['出産祝い', 5000], ['結婚祝い', 5000], ['御霊前', 5000], ['お見舞い', 3000], ['出産祝い', 5000]]

#[2]辞書　同一のkeyの値は累計額を加算する　結婚お祝いなら、10,000+5000=15,000
#dicのキーが存在しない場合　if key not in dic.keys() 他にdic.get(key)
#dicの値を取り出すのは　dic[key] dic["結婚お祝い"] + spend_list[c1]
v = 0
for c in spend_list:
    spend_koumoku = c[0]
    #辞書のキーが存在しない場合は、キーを追加
    if spend_koumoku not in dic_spend.keys():
        dic_spend[spend_koumoku] = c[1]
    else:
        #辞書のキーが存在する場合は、アイテムを加算
        dic_spend[spend_koumoku] = dic_spend[spend_koumoku] + c[1]

#[3] 辞書のキーと値を取り出すのは、 for k,v in dic.items()
for c,v in dic_spend.items():
    print(c,v)
#実行結果　結婚祝い 15000　お見舞い袋 100　出産祝い 10000　御霊前 5000　お見舞い 3000

#[4] 辞書型（dictionary）の値をvaluesでリスト型にして、まとめて取り出す dic.values()　
# 実行結果dict_values([15000, 100, 10000, 5000, 3000])
#　辞書のアイテムをsum関数で足すには　sum(dic.value()のリスト)
dic_val = dic_spend.values()
print(dic_val)
print(sum(dic_val))

#[5] 総覧表　D,E列に書き出し まずはリストに格納
r_list = []
for row in sh["B4:C9"]:
    ls =[]
    for r in row:
        ls.append(r.value)
    r_list.append(ls)
print(r_list)

#[6] 支払総計表への転記
# リストの値 = リスト型　class List
# リストの要素を取り出す　str型　リスト[0] = str
# 辞書のvalueを比較する場合、同じstr型出ないと一致しない　よって、リスト[0]（添え字）と辞書の値を比較する
# 支払累計額のリストを、ruikei_listで作成　最後に総類額をsum(ruikei_list.values())で取り出して転記

ruikei_list = []
for cnt,val in enumerate(r_list):
    print(val,val[0],sep="*")
    koumoku = val[0]
#リストの要素が、4月支払リストの辞書のキーにない場合は、先月までの支払総額を、支払累計額に転記する
#   continueでそれ以降を中断　ループを続行
#リストの要素が、4月支払リストの辞書のキーにある場合は、4月のKeyの値を、今月支払額に転記する
#   今月支払額は、4月支払リストの辞書の値、支払累計学は、sum([先月までの支払総額,4月支出の辞書の値])
#   breakでループを終了
    if val[0] not in dic_spend:
        print("ありません",val[0])
        sh.cell(row=4 + cnt, column=5).value = val[1]
        ruikei_list.append(val[1])
        continue
    else:
        for c,v in dic_spend.items():
            if koumoku == c:
                sh.cell(row=4+cnt,column=4).value = v
                sh.cell(row=4+cnt,column=5).value = sum([val[1],v])
                ruikei_list.append(sum([val[1],v]))
                print(v,"v",sum([val[1],v]),sep=":")
                break

#[7]今月支払額の累計　sum(4月支払リストの辞書.values())
#   辞書型（dictionary）の値をvaluesでリスト型にして、まとめて取り出す sum(list名)
#   支払累計額 リストの累計額は、sum([list])
sh.cell(row=10,column=4).value = sum(dic_spend.values())
sh.cell(row=10,column=5).value = sum(ruikei_list)
wb.save(mod_ph)

