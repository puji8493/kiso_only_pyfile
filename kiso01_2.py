#[1]シート「mondai1」で、各行につき、「合計」の値が100を超えていたらA列に「○」そうでなければ「×」をつけるマクロを作りなさい
#シート「mondai2」で、[1]～[3]と同様の作業をするマクロを作りなさいただし、「○」や「×」を記入する先は、J列とする

# kiso01_judge.pyのover_time_judge関数を呼び出す工夫をした

import openpyxl
import kiso01_Judge
from openpyxl.styles import Font

wbPath = "../5bai/kiso001_kaito.xlsx"
svPath = "../5bai/kiso001_kaito_mod_2.xlsx"
wb = openpyxl.load_workbook(wbPath)
ws = wb["Mondai2"]

r_idx = 4
for r in ws["B4:J13"]:#J列へ書き込むため、J列のインデックス番号を取得するため

    # kiso01_judge.pyのover_time_judge関数を呼び出す
    ans = kiso01_Judge.over_time_judge(r[7].value)

    #ws.cell(row=r_idx,column=10).value = ans
    r[8].value =ans
    r[8].font = Font(bold=True,italic=True)
    r_idx = r_idx + 1

wb.save(svPath)