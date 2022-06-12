#フリガナ10文字だと”長すぎ" 所属部署に”都島”があると、それを表示する　リストのスライスにした
#"都島"の存在確認は、startswithメソッド　syozoku.startswith("都島")でもいい

import openpyxl
ph = "../learning/ks014.xlsx"
mod_ph = "../learning/ks014_mod.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["IfThenEndIf"]

#J列ひとことにコメントする戻り値を求める関数
#"都島"の存在確認は、startswithメソッド　syozoku.startswith("都島")でもいい

def Hitokoto(syozoku,furigana):
    if syozoku[:2] == "都島" and len(furigana) >= 10:
        coment = "都島グループです フリガナが長すぎます"
    elif syozoku[:2] == "都島":
        coment = "都島グループです"
    elif len(furigana) >= 10:
        coment = "フリガナが長すぎます"
    else:
        coment =""
    return coment

#２行目からシートのデータをリストに格納
ls = []
for row in sh.iter_rows(min_row=2,max_row=11):
    ls.append(row)

for c,v in enumerate(ls):
    print(v[0].value)


for row in sh["A2:D10"]:
    id_list=[]
    for c in row:
        id_list.append(c.value)
        #print(id_list)
#print(id_list)


#リストの値とインデックス番号を書き出す
for c,v in enumerate(ls):
    #所属はB列　フリガナはD列（インデックス番号3)
    syozoku = v[1].value
    furigana = v[3].value
    ans = Hitokoto(syozoku,furigana)
    #print(ans,c)
    sh.cell(row=c+2,column=10).value = ans


wb.save(mod_ph)