#残業時間が100時間を超えているを要注意リストに格納して、書き出す
#計算式をCellオブジェクトから取得したい時は、引数　data_only=Trueを追加する

import openpyxl

ph = "../learning/ks005.xlsx"
mod_ph = "../learning/ks005_mod.xlsx"
#計算式をCellオブジェクトから取得したい時は、引数　data_only=Trueを追加する
wb = openpyxl.load_workbook(ph,data_only=True)
shFm = wb["元データ"]
shTo = wb["要注意リスト"]

#残業リストを作る
zangyo_lst = []
for row in shFm.iter_rows(min_row=4,max_row=13,max_col=9):
    #行の空リストをつくる
    ls = []
    for r in row:
        ls.append(r.value)
    zangyo_lst.append(ls)

#合計列が１００時間以上の場合は、要注意リストに格納する
a_list = [t for t in zangyo_lst if t[8] >= 100]
print(a_list)

#sheets"要注意リスト"に、100時間残業敬の人を書き出す
#リストの合計時間はI列（インデックス番号8)
for c,v in enumerate(a_list):
    shTo.cell(row=c+9,column=3).value = v[0]
    shTo.cell(row=c+9,column=4).value = v[1]
    shTo.cell(row=c+9,column=5).value = v[8]

wb.save(mod_ph)
