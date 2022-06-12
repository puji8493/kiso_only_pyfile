#点数により判定を返す
#関数hanteiで戻り値を返す
import openpyxl

ph = "../learning/ks003_kaito.xlsx"
mod_path = "../learning/ks003_kaito_mod_2.xlsx"

wb = openpyxl.load_workbook(ph)
sh = wb["Sheet1"]

def Hantei(tokuten):
    if tokuten > 80:
        gouhi = "A判定です"
    elif tokuten > 60:
        gouhi = "B判定です"
    elif tokuten > 40:
        gouhi = "C判定です"
    else:
        gouhi = "D判定です"
    return gouhi

#行の要素を保存するリスト
gyo_list = []
for row in sh.iter_rows(min_row=2,max_row=11,max_col=4):
    #行を読み込むリスト
    lst = []
    #行の中をループして、いったんリストに保存
    for r in row:
        lst.append(r.value)
    #print(lst)#実行結果 [1, '松本 めぐみ', '女性', 65]
    #gyo_listに２９行目のlstのリストを保存する
    gyo_list.append(lst)
    print(gyo_list)
    #実行結果
    #[[1, '松本 めぐみ', '女性', 65]]
    #[[1, '松本 めぐみ', '女性', 65], [2, '金沢 京子', '女性', 75]]
    #[[1, '松本 めぐみ', '女性', 65], [2, '金沢 京子', '女性', 75], [3, '東京 尚広', '男性', 82]]

#gyo_listをenumerate関数で、インデックス番号と値を取り出す
for c,v in enumerate(gyo_list):
    #gyo_listのインデックス番号3 = 得点を関数の引数に渡す
    ans = Hantei(v[3])
    sh.cell(row=c+2,column=5).value = ans

wb.save(mod_path)