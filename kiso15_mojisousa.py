#路線駅名 小田急線/代々木八幡駅/歩4といった文字列の集合から、 "/"の区切り文字をsplitメソッドで分割してセルに書き出す
#建物「RC」は「鉄筋コンクリート」、「SRC」は「鉄骨鉄筋コンクリート」に変換する

import openpyxl
ph = "../learning/ks015.xlsx"
mod_ph = "../learning/ks015_mod.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["問題"]

route_list = []
for row in sh.iter_rows(min_row=3,max_col=5):
    ls = []
    for r in row:
        ls.append(r.value)
    route_list.append(ls)

for c,v in enumerate(route_list):

    #路線駅名はリスト要素4番目
    rosen = v[4]
    #　splitメソッド
    sp = rosen.split("/")
    # splitメソッドの戻り値（list)を、要素の数だけ転記する　たとえば、4行目/一つしかないなら、2列の出力
    for cnt,val in enumerate(sp):
        moji = val
        sh.cell(row=c+3,column=6+cnt).value = moji

    #建物タイプ　表示ルールは、タイプ/階数/何階建と"/"が二つある前提
    kouzou = v[3]
    sp_k = kouzou.split("/")
    blt = sp_k[0]
    if blt == "RC":
        built_type = "鉄筋コンクリート"
    elif blt == "SRC":
        built_type = "鉄骨鉄筋コンクリート"
    else:
        built_type = "リスト外"
    floors = sp_k[1] + "階"
    story = sp_k[2].replace("F","階建て")

    sh.cell(row=c+3, column=9).value = built_type
    sh.cell(row=c+3, column=10).value = floors
    sh.cell(row=c+3, column=11).value = story

    #formatメソッドで説明文を作る
    f = f"この建物は、{built_type}で、{story}の{floors}部分です"
    sh.cell(row=c+3, column=12).value = f

    print(f)
    # 実行結果　この建物は鉄筋コンクリートで、7F建ての2階部分です

    #findメソッド
    print(f.find("7"))
    #実行結果　この建物は、鉄筋コンクリートで、7階建ての6階部分です　→　16

    #文字を2飛ばしづつに出力
    print(f[::2])
    #実行結果　この建物は、鉄筋コンクリートで、3階建ての3階部分です →　こ建は鉄コクーで3建の階分す


wb.save(mod_ph)