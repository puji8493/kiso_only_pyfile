#和暦を西暦にする
#男性の場合は80点以上、女性の場合は70点以上を合格して、合否を書き出す
#for文でシートの行を読み込むとき、iter_rows()メソッドで範囲を指定できる。例えばmin_row=２行目とかmax_column8列目までとか
#row_listリストの値を書き出す enumerate関数でリストのインデックス番号を取得、書き出すセルの行番地の変数に代入する
#行の開始行は２行目のため、c + 2 と２を加算する

import openpyxl
fPath = ("../learning/ks002.xlsx")
mod_Path = ("../learning/ks002_mod_2.xlsx")
wb = openpyxl.load_workbook(fPath)
sh = wb["Sheet1"]

def hantei(gender,tokuten):
    if gender == "男性" and tokuten >= 80:
        gouhi = "合格"
    elif gender == "女性" and tokuten >= 70:
        gouhi = "合格"
    else:
        gouhi =""
    return gouhi

def trans_seireki(gengo,birth_year):
    if gengo == "昭和":
        ad = 1925 + birth_year
    else:
        ad = 1988 + birth_year
    return ad
#1行ごとの要素を保存するリスト [２行目ID,氏名,性別・・][3行目ID,氏名,性別・・]
row_list = []
#iter_rows()メソッドで範囲を指定してループ2行目から11行目　1列目から8列目まで
for row in sh.iter_rows(min_row=2,max_row=11,min_col=1,max_col=6):
    #空リスト
    lst = []
    for c in row:
        lst.append(c.value)
    #上のlstのリストを、row_listに入れる
    row_list.append(lst)
    #printした時の実行結果[1, '松本 めぐみ', '女性', 65, '昭和', 31][2, '金沢 京子', '女性', 75, '昭和', 63]

#row_listリストの値を書き出す enumerate関数でリストのインデックス番号を取得、書き出すセルの行番地の変数に代入する
#行の開始行は２行目のため、c + 2 と２を加算する
for c,v in enumerate(row_list):
    #v[4]生年元号とv[5]生年の値を関数trans_seirekiの引数として渡す
    seireki = trans_seireki(v[4],v[5])
    #v[2]性別とv[3]得点を、関数hanteiの引数として渡す
    gouhi_hantei = hantei(v[2],v[3])
    sh.cell(row=c+2,column=7).value = seireki
    sh.cell(row=c+2,column=8).value = gouhi_hantei

wb.save(mod_Path)
