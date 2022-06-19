#当月の支出リストに格納した後、総元帳に転記する　最終行に合計額を転記する
#当月（4月）の支出項目を辞書で登録　キーが重複した時は、値を加算する　連想配列で値の累計を出すイメージ

import openpyxl
ph = "../learning/ks011.xlsx"
mod_ph = "../learning/ks011_mod_dic2_2.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet1"]

#4月の支出項目を辞書で登録　キーが重複した時は、値を加算する　連想配列で値の累計を出すイメージ
lst_Aplil = {}
for row in sh["I4:J10"]:
    syousai = row[0].value
    #辞書のキーがすでに存在していたら、valueに金額を加算する
    #新規のキーの時は、金額をvalueに代入する
    if lst_Aplil.get(syousai):
        v = lst_Aplil[syousai] + row[1].value
        lst_Aplil[syousai] = v
    else:
        lst_Aplil[syousai] = row[1].value

#今月支払と支払累計の変数をつくり、値を転記した金額の累計を最後に出力する
pay_thismonth = 0
pay_total = 0

#Ｂ列の項目と一致したら、4月支出金額の値をＤ列今月支払額に転記する
#A列からC列　総覧表をリストに格納
souran_list = []
for row in sh.iter_rows(min_row=4,max_row=9,max_col=3):
    #行をリストに格納
    ls = []
    for r in row:
        ls.append(r.value)
    #総覧リストに格納
    souran_list.append(ls)

for c,i in enumerate(souran_list):
    koumoku = i[1]
    #項目が辞書に存在しない場合は、総計に転記し、総支払額を追加してbreak文でループを抜ける
    #項目じ辞書に存在する場合は、今月支出額に転記し、今月累計額を加算してbreak文でループを抜ける
    for k,v in lst_Aplil.items():
        if koumoku not in lst_Aplil:
            sh.cell(row=c+4,column=5).value = i[2]
            pay_total = pay_total + sh.cell(row=c+4,column=5).value
            break
        elif k == koumoku:
            sh.cell(row=c+4,column=4).value = v
            sh.cell(row=c+4,column=5).value = v + sh.cell(row=c+4,column=3).value
            pay_thismonth = pay_thismonth + sh.cell(row=c+4,column=4).value
            pay_total = pay_total + sh.cell(row=c+4,column=5).value
            break
        else:
            print("不要なループ")
            continue

sh["D10"].value = pay_thismonth
sh["E10"].value = pay_total
wb.save(mod_ph)

