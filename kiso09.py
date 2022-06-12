#苗字と名前を split()メソッド　引数"/" もしくは"_"の文字列で分割したリストを作り、C/D列に書き出し　
#sheets"問題3” 区切り文字は"/"と" "が混在
import openpyxl
ph = "../learning/ks009.xlsx"
mod_ph = "../learning/ks009_mod3.xlsx"
wb = openpyxl.load_workbook(ph)
sh3 = wb["問題３"]

def get_first_last_name(namae):
    sp_list = []
    if "/" in namae:
        sp_list = namae.split("/")
    elif " " in namae:
        sp_list = namae.split(" ")

    return sp_list
    #print(sp_list[0],sp_list[1],sep=":")

#名簿全員のリストを格納　（1列のみ）
members_list = []
for row in sh3.iter_rows(min_row=2,max_row=11,min_col=2,max_col=2):
    #行の空リスト
    ls = []
    #行の要素の値をmembers_listに格納
    for r in row:
        ls.append(r.value)
    members_list.append(ls)

#名前を文字列から苗字・氏名をsplitメソッドで分割する関数に渡し、戻り値を所定のセルに書き出す
#シートに書き出す行番号のカウンターcnt
cnt = 2
for c in members_list:
    #名前の変数をmembers_list[0]に格納する
    namae = c[0]
    #関数に名前の変数を渡し、bufという変数名のリスト格納する
    buf = []
    buf = get_first_last_name(namae)
    #split関数の戻り値のタイプはリスト
    print(type(buf),buf[0],buf[1],sep=":")
    sh3.cell(row=cnt,column=3).value = buf[0]
    sh3.cell(row=cnt,column=4).value = buf[1]
    cnt = cnt +1

wb.save(mod_ph)
