#キャンペーン名簿表を元に、転記リストと一致したものを書き出す
#キャンペーン応募状況のIDは一意＝重複しないので、辞書でkey=ID value=キャンペーンタイプの値を取得する
import openpyxl
ph = "../learning/ks013.xlsx"
mod_ph = "../learning/ks013_mod.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["キャンペーン名簿"]

#キャンペン応募状況の辞書を作る　key=ID value=キャンペンタイプ
dic = {}
for row in sh["E11:F22"]:
    ind = row[0].value
    dic[ind] = row[1].value

#Ａ列のＩＤと、キャンペン応募状況の辞書のkey(ID)が一致したらＣ列に書き出す
#辞書の値と一致、セルへ書き出したらbreak文でfor文を抜ける

kokyaku_list = []
for row in sh.iter_rows(min_row=4,max_row=29,max_col=2):
    ls = []
    for r in row:
        ls.append(r.value)
    kokyaku_list.append(ls)

for c,val in kokyaku_list:
    print(c,val,c,sep=":")
    for k,v in dic.items():
        if c == k:
            print(k,v,c)
            sh.cell(row=c+3,column=3).value = v
            break

wb.save(mod_ph)