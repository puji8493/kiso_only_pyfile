import csv
import openpyxl

date_type_list = []
with open("test_data_microsecond.csv",mode="r",encoding="utf-8") as f:
    for row in csv.DictReader(f):
        date_type_list.append(row)

dic = {}
for val in date_type_list:
    key = val["date"]
    tmp_dic = {}
    for k,v in val.items():
        if k != "date":
            tmp_dic[k] = v
    print(tmp_dic)
    dic[key] = tmp_dic

for k,v in dic.items():
    print(k,v,sep="☆")

start_time = "2022-03-01 10:00:00.001000"
end_time = "2022-03-01 10:00:00.031000"

#転記先の見出行のリスト
wb = openpyxl.load_workbook("test_data_microsecond.xlsx")
wb_mod = "test_data_microsecond_mod.xlsx"
sh = wb["data"]
items_list = []
for row in sh.iter_rows(min_row=1,min_col=2,max_row=1,max_col=4):
    for r in row:
        items_list.append(r.value)
print(items_list)

cnt = 2
#key:date v:itemsとdataのvalue
for d,val in dic.items():
    if start_time < d < end_time:
        sh.cell(row=cnt,column=1).value = d
        print(val)
        for k,v in val.items():
            for c, i in enumerate(items_list):
                if k == i:
                    sh.cell(row=cnt,column=c+2).value = v
        cnt += 1

wb.save(wb_mod)
