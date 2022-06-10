#[2]シート「Sheet2」で、1月の残業時間が一番多い人の氏名をセルK4に、「1月」という文字列をセルL4に、残業時間の値をセルM4に記入するマクロを作りなさい。ただし、残業時間が一番多い人はひとりしかいないものとしてよい）
#一番得点の高い人を重複させたデータをつくる
#最大値が２つ以上あった場合を考慮し、最大値のインデックス番号をリストに格納する
#最大値が何回目のループ回数か調べたいので、enumerate関数を使ってインデックスを抽出　#抽出したインデックス番号をlist.appned()メソッドで格納。appendメソッドつけ忘れて取り扱えないオブジェクトエラー注意

import openpyxl

ph = "../5bai/ks012.xlsx"
mod_ph = "../5bai/ks012_mod2.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet2"]

#名前と1月残業時間をリストに格納
ls = []
#残業時間を格納するリスト
value_ls = []

#名前、残業時間を格納するリスト、残業時間のみmax関数で格納するリストをつくる
for c in sh["B6:C33"]:
    ls.append([c[0].value,c[1].value])
    value_ls.append(c[1].value)
    print(ls)

#最大値をmax関数で格納
max_value = max(value_ls)
print(max_value)

#最大値と一致する場合は、セルK4以降名前、月、値を書き出す
cnt = 4
for c,v in enumerate(ls):
    if max_value == v[1] :
       sh.cell(row=cnt,column=11).value = v[0]
       sh.cell(row=cnt,column=12).value = "１月"
       sh.cell(row=cnt,column=13).value = v[1]
       cnt = cnt + 1
wb.save(mod_ph)