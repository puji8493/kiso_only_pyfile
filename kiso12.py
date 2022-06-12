#一番高い得点の列に”最高点です”と書き出し
#最高点が重複したら、全員書き出す
#最大値が２つ以上あった場合を考慮し、最大値のインデックス番号をリストに格納する
#最大値が何回目のループ回数か調べたいので、enumerate関数を使ってインデックスを抽出　#抽出したインデックス番号をlist.appned()メソッドで格納。appendメソッドつけ忘れて取り扱えないオブジェクトエラー注意

import openpyxl
ph = "../learning/ks012.xlsx"
mod_ph = "../learning/ks012_mod.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet1"]

tokuten_list = []
for row in sh.iter_rows(min_row=2,max_row=11,max_col=10):
    tokuten_list.append(row[6].value)

#最大値の値とインデックスを取得し、Ｊ列に転記する
# max関数（引数リスト）で、最大値を求められる
# list.index（引数今回は最大値を代入した変数）で、リストのインデックス番号を調べられる
max_value = max(tokuten_list)

#最大値が２つ以上あった場合を考慮し、最大値のインデックス番号をリストに格納する
#最大値が何回目のループ回数か調べたいので、enumerate関数を使ってインデックスを抽出
#抽出したインデックス番号をlist.appned()メソッドで格納。appendメソッドつけ忘れて取り扱えないオブジェクトエラー注意

max_gyo_list = []
for c,v in enumerate(tokuten_list):
    if max_value == v:
        max_gyo_list.append(c)
        #print(max_gyo_list,"max",sep=":")

#最高点のリストの行に、”最高点です”とフラグをたてる
for i in max_gyo_list:
    sh.cell(row=i+2,column=10).value = "最高点☆です"

wb.save(mod_ph)

