#[2]シート「Sheet2」で、1月の残業時間が一番多い人の氏名をセルK4に、「1月」という文字列をセルL4に、残業時間の値をセルM4に記入するマクロを作りなさい。ただし、残業時間が一番多い人はひとりしかいないものとしてよい）
#一番得点の高い人はを重複させたデータ　40.5　二名　40.0　1名いるとする
#最大値が２つ以上あった場合を考慮し、最大値のインデックス番号をリストに格納する
#最大値が何回目のループ回数か調べたいので、enumerate関数を使ってインデックスを抽出　#抽出したインデックス番号をlist.appned()メソッドで格納。appendメソッドつけ忘れて取り扱えないオブジェクトエラー注意

#最大値をfloat型で返すには、numpyを利用する
#インストールはコマンドプロンプトでpip install numpy　ＤＬしてから、import numpy 分かりやすい参考リンクhttps://udemy.benesse.co.jp/data-science/ai/python-numpy.html
# arry の公式ドキュメント　https://docs.python.org/ja/3/library/array.html
#numpyの配列の取り扱い　分かりやすかったリンク　https://python.atelierkobato.com/max/

import openpyxl
import numpy as np

ph = "../5bai/ks012.xlsx"
mod_ph = "../5bai/ks012_mod2.xlsx"
wb = openpyxl.load_workbook(ph)
sh = wb["Sheet2"]

#[2]シート「Sheet2」で、1月の残業時間が一番多い人の氏名をセルK4に、「1月」という文字列をセルL4に、残業時間の値をセルM4に記入するマクロを作りなさい。ただし、残業時間が一番多い人はひとりしかいないものとしてよい）
#名前＋月＋残業時間をリストに個別に格納し、一番大きい残業時間を調べる
#1行毎にTupleで格納されるので、改めてList（残業リスト）に格納する　list.append([  ])
zangyo_list = []
jikan_list= []
for row in sh["B6:C33"]:
    zangyo_list.append([row[0].value,row[1].value])
    jikan_list.append(row[1].value)

#時間のリストをタプル関数でタプルにしてから、NumPy配列（np.array）の要素の中から最大値をfloat型で取得する
jikan_tuple = tuple(jikan_list)
jikan_array = np.array(jikan_tuple)

#numpy.nanmax() は nan を無視して最大要素を float型で返します。
mx = np.nanmax(jikan_array)
print(mx)
#numpy.argmax() は配列の最大要素のインデックスを返します。
ind = np.argmax(jikan_array)
print(ind)

#最大値（float型）から最大値のインデックスを取得する　重複した場合もあるので、リストに格納
ind_list = []
for c,v in enumerate(jikan_list):
    if v == mx:
        print(c,"☆",sep=":")
        ind_list.append(c)
print(ind_list)

#最大値のインデックス番号をもつ残業リストを転記
for c,v in enumerate(ind_list):
    sh.cell(row=4+c,column=11).value= zangyo_list[v][0] #名前
    sh.cell(row=4+c,column=12).value= zangyo_list[v][1] #時間


wb.save(mod_ph)

#これは次回
#[3]シート「Sheet2」で、1～6月の残業時間が一番多い人の氏名をセルK4に、月名をセルL4に、残業時間の値をセルM4に記入するマクロを作りなさい。（ただし、残業時間が一番多い人はひとりしかいないものとしてよい）
